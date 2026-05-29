#!/usr/bin/env python3
"""Combineer twee Kwandoo-aanwezigheidslijsten (Jadadde + Villa Max) tot één Excel.

Gebruik:
    python3 combineer_aanwezigheidslijsten.py LIJST1.pdf LIJST2.pdf [-o uitvoer.xlsx]

De volgorde van de twee PDF's maakt niet uit: het script herkent zelf welke
de Jadadde- (speelplein) en welke de Villa Max- (opvang) lijst is op basis van
de inhoud. Datums en dagkolommen worden uit de PDF gelezen.

Resultaat: een Excel met 3 tabbladen
  1. Gecombineerd (overlap)  - kinderen op BEIDE lijsten
  2. Alleen Jadadde          - enkel speelplein
  3. Alleen Villa Max        - enkel opvang
"""

import argparse
import os
import re
import sys
import unicodedata
from datetime import date

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAANDEN_NL = ["", "januari", "februari", "maart", "april", "mei", "juni",
              "juli", "augustus", "september", "oktober", "november", "december"]

# ----------------------------------------------------------------------------
# PDF lezen
# ----------------------------------------------------------------------------

def lees_pdf_tekst(pdf_path):
    """Geef alle tekst van de PDF terug als één string (per regel)."""
    if not os.path.isfile(pdf_path):
        sys.exit(f"FOUT: bestand niet gevonden: {pdf_path}")
    delen = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tekst = page.extract_text()
                if tekst:
                    delen.append(tekst)
    except Exception as e:  # corrupte / niet-leesbare PDF
        sys.exit(f"FOUT: kan PDF niet lezen ({pdf_path}): {e}")
    return "\n".join(delen)


def detecteer_type(tekst):
    """Bepaal of de PDF een Jadadde- of Villa Max-lijst is."""
    lower = tekst.lower()
    if "villa max" in lower:
        return "villamax"
    if "jadadde" in lower:
        return "jadadde"
    return None


def extract_namen(tekst):
    """Haal de kindernamen uit de tekst.

    Een naamregel begint met het rijnummer gevolgd door minstens één woord
    met een letter. Lege genummerde rijen en kop-/voetregels vallen weg.
    Dedupliceert op matchsleutel (eerste schrijfwijze blijft behouden).
    """
    gezien = set()
    namen = []
    for regel in tekst.split("\n"):
        m = re.match(r"^\s*\d+\s+(.+)$", regel)
        if not m:
            continue
        naam = m.group(1).strip()
        # Moet minstens één letter bevatten en geen kopregel zijn.
        if not re.search(r"[A-Za-zÀ-ÿ]", naam):
            continue
        if naam.lower().startswith(("aanwezigheids", "naam en voornaam")):
            continue
        sleutel = match_sleutel(naam)
        if not sleutel or sleutel in gezien:
            continue
        gezien.add(sleutel)
        namen.append(naam)
    return namen


def parse_datums(tekst):
    """Lees de datums uit de 'Datums:'-regel. Geeft een lijst datetime.date terug."""
    m = re.search(r"Datums?\s*:\s*([0-9/,\s]+)", tekst)
    datums = []
    if m:
        for stuk in re.findall(r"(\d{1,2})/(\d{1,2})/(\d{4})", m.group(1)):
            d, mnd, jr = map(int, stuk)
            datums.append(date(jr, mnd, d))
    return sorted(set(datums))


# ----------------------------------------------------------------------------
# Naam normaliseren / matchen
# ----------------------------------------------------------------------------

def _strip_accenten(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c))


def _schoon(naam):
    """Verwijder ruis: tekst tussen haakjes (bv. '(papa)'), sterretjes, accenten."""
    naam = re.sub(r"\([^)]*\)", " ", naam)   # (papa) e.d.
    naam = naam.replace("*", " ")
    naam = _strip_accenten(naam.lower())
    naam = re.sub(r"[^a-z\s]", " ", naam)
    return re.sub(r"\s+", " ", naam).strip()


def match_sleutel(naam):
    """Volgorde-onafhankelijke sleutel: gesorteerde, geschoonde tokens.

    Hierdoor matchen 'Bruwier Arthur' en 'Arthur Bruwier', en valt ruis als
    '(papa)' of accentverschillen ('Téo' vs 'Teo') weg.
    """
    return " ".join(sorted(_schoon(naam).split()))


def sorteer_sleutel(naam):
    """Consistente weergavevolgorde over alle tabbladen: op achternaam (= eerste
    token in 'Achternaam Voornaam'), geschoond en accentvrij."""
    return _schoon(naam)


# ----------------------------------------------------------------------------
# Datum-helpers voor titels en kolomlabels
# ----------------------------------------------------------------------------

def dagdeel_labels(datums):
    """['1/7 VM', '1/7 NM', ...] voor elke datum."""
    labels = []
    for d in datums:
        labels.append(f"{d.day}/{d.month} VM")
        labels.append(f"{d.day}/{d.month} NM")
    return labels


def periode_tekst(datums):
    if not datums:
        return ""
    d0, dn = datums[0], datums[-1]
    if d0 == dn:
        return f"{d0.day} {MAANDEN_NL[d0.month]} {d0.year}"
    if d0.month == dn.month and d0.year == dn.year:
        return f"{d0.day}-{dn.day} {MAANDEN_NL[d0.month]} {d0.year}"
    return f"{d0.day}/{d0.month} - {dn.day}/{dn.month}/{dn.year}"


# ----------------------------------------------------------------------------
# Excel-opmaak
# ----------------------------------------------------------------------------

BLAUW_DARK = "1F4E79"
BLAUW_MID = "2E75B6"
GROEN_DARK = "375623"
GROEN_MID = "548235"
GRIJS_LIGHT = "F2F2F2"
WIT = "FFFFFF"

_thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def header_cell(ws, row, col, value, bg, fg="FFFFFF", wrap=True, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=fg, name="Arial", size=9)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=wrap)
    c.border = BORDER
    return c


def data_cell(ws, row, col, value="", bg=WIT, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=9)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center")
    c.border = BORDER
    return c


def vul_rijen(ws, namen, eerste_rij, laatste_kol):
    """Schrijf genummerde naamrijen met afwisselende achtergrond."""
    for i, naam in enumerate(sorted(namen, key=sorteer_sleutel), start=1):
        row = eerste_rij + i - 1
        bg = GRIJS_LIGHT if i % 2 == 0 else WIT
        data_cell(ws, row, 1, i, bg)
        data_cell(ws, row, 2, naam, bg, center=False)
        for col in range(3, laatste_kol + 1):
            data_cell(ws, row, col, "", bg)
        ws.row_dimensions[row].height = 18


# ----------------------------------------------------------------------------
# Tabbladen bouwen
# ----------------------------------------------------------------------------

def bouw_tab_gecombineerd(wb, namen, datums):
    """Tab 1: kinderen op beide lijsten, kolommen van Jadadde én Villa Max."""
    daglabels = dagdeel_labels(datums)
    n = len(daglabels)  # aantal dagdeel-kolommen per activiteit

    ws = wb.active
    ws.title = "Gecombineerd (overlap)"

    # Kolomindeling dynamisch opbouwen.
    col = 1
    c_num = col; col += 1
    c_naam = col; col += 1
    j_in = col; j_uit = col + 1; col += 2
    j_opv1 = col; j_opv2 = col + 1; col += 2
    j_dag0 = col; col += n
    v_warm = col; v_geen = col + 1; v_niet = col + 2; col += 3
    v_dag0 = col; col += n
    c_opm = col
    laatste = c_opm

    laatste_letter = get_column_letter(laatste)
    periode = periode_tekst(datums)

    ws.merge_cells(f"A1:{laatste_letter}1")
    t = ws["A1"]
    t.value = f"Gecombineerde aanwezigheidslijst — Jadadde + Villa Max ({periode})"
    t.font = Font(bold=True, color=WIT, name="Arial", size=12)
    t.fill = PatternFill("solid", start_color=BLAUW_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{laatste_letter}2")
    leg = ws["A2"]
    leg.value = "Enkel kinderen die op BEIDE lijsten staan. Vul in/uit tijden ter plaatse in."
    leg.font = Font(italic=True, name="Arial", size=9, color="595959")
    leg.alignment = Alignment(horizontal="center", vertical="center")
    leg.fill = PatternFill("solid", start_color=GRIJS_LIGHT)

    def mc(c0, c1, row=3):
        ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")

    # Rij 3 = groepen, rij 4 = subkolommen.
    ws.merge_cells(f"{get_column_letter(c_num)}3:{get_column_letter(c_num)}4")
    header_cell(ws, 3, c_num, "#", BLAUW_DARK)
    ws.merge_cells(f"{get_column_letter(c_naam)}3:{get_column_letter(c_naam)}4")
    header_cell(ws, 3, c_naam, "Naam en Voornaam", BLAUW_DARK)

    mc(j_in, j_uit); header_cell(ws, 3, j_in, "Tijdregistratie Jadadde", BLAUW_MID)
    header_cell(ws, 4, j_in, "In", BLAUW_MID)
    header_cell(ws, 4, j_uit, "Uit", BLAUW_MID)

    mc(j_opv1, j_opv2); header_cell(ws, 3, j_opv1, "Opvang Jadadde", BLAUW_MID)
    header_cell(ws, 4, j_opv1, "Ochtend/Avond", BLAUW_MID)
    header_cell(ws, 4, j_opv2, "Middag", BLAUW_MID)

    mc(j_dag0, j_dag0 + n - 1)
    header_cell(ws, 3, j_dag0, "JADADDE — Aanwezigheid per dagdeel", BLAUW_MID)
    for k, lbl in enumerate(daglabels):
        header_cell(ws, 4, j_dag0 + k, lbl, BLAUW_MID)

    mc(v_warm, v_niet)
    header_cell(ws, 3, v_warm, "Maaltijd / Speelplein (Villa Max-keuze)", GROEN_MID)
    header_cell(ws, 4, v_warm, "Warme maaltijd", GROEN_MID)
    header_cell(ws, 4, v_geen, "Geen maaltijd", GROEN_MID)
    header_cell(ws, 4, v_niet, "Niet speelplein", GROEN_MID)

    mc(v_dag0, v_dag0 + n - 1)
    header_cell(ws, 3, v_dag0, "VILLA MAX — Aanwezigheid per dagdeel", GROEN_MID)
    for k, lbl in enumerate(daglabels):
        header_cell(ws, 4, v_dag0 + k, lbl, GROEN_MID)

    ws.merge_cells(f"{get_column_letter(c_opm)}3:{get_column_letter(c_opm)}4")
    header_cell(ws, 3, c_opm, "Opmerking", BLAUW_DARK)

    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 28

    vul_rijen(ws, namen, eerste_rij=5, laatste_kol=laatste)

    # Kolombreedtes.
    ws.column_dimensions[get_column_letter(c_num)].width = 4
    ws.column_dimensions[get_column_letter(c_naam)].width = 24
    for c in (j_in, j_uit, j_opv1, j_opv2):
        ws.column_dimensions[get_column_letter(c)].width = 8
    for c in range(j_dag0, j_dag0 + n):
        ws.column_dimensions[get_column_letter(c)].width = 7
    for c in (v_warm, v_geen, v_niet):
        ws.column_dimensions[get_column_letter(c)].width = 10
    for c in range(v_dag0, v_dag0 + n):
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.column_dimensions[get_column_letter(c_opm)].width = 18


def bouw_tab_alleen(wb, titel, namen, datums, kleur_dark, kleur_mid, met_maaltijd):
    """Tab 2/3: kinderen op slechts één lijst."""
    daglabels = dagdeel_labels(datums)
    n = len(daglabels)

    ws = wb.create_sheet(titel)

    col = 1
    c_num = col; col += 1
    c_naam = col; col += 1
    if met_maaltijd:
        v_warm = col; v_geen = col + 1; v_niet = col + 2; col += 3
        extra_groep = ("Maaltijd / Speelplein",
                       [(v_warm, "Warme maaltijd"), (v_geen, "Geen maaltijd"),
                        (v_niet, "Niet speelplein")])
    else:
        e_in = col; e_uit = col + 1; col += 2
        e_opv1 = col; e_opv2 = col + 1; col += 2
        extra_groep = None
    dag0 = col; col += n
    c_opm = col
    laatste = c_opm

    laatste_letter = get_column_letter(laatste)
    periode = periode_tekst(datums)

    ws.merge_cells(f"A1:{laatste_letter}1")
    t = ws["A1"]
    t.value = titel + f" ({periode})"
    t.font = Font(bold=True, color=WIT, name="Arial", size=11)
    t.fill = PatternFill("solid", start_color=kleur_dark)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    def mc(c0, c1, row=2):
        ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")

    ws.merge_cells(f"{get_column_letter(c_num)}2:{get_column_letter(c_num)}3")
    header_cell(ws, 2, c_num, "#", kleur_dark)
    ws.merge_cells(f"{get_column_letter(c_naam)}2:{get_column_letter(c_naam)}3")
    header_cell(ws, 2, c_naam, "Naam en Voornaam", kleur_dark)

    if met_maaltijd:
        v_warm, v_geen, v_niet = (extra_groep[1][0][0], extra_groep[1][1][0],
                                  extra_groep[1][2][0])
        mc(v_warm, v_niet)
        header_cell(ws, 2, v_warm, "Maaltijd / Speelplein", kleur_mid)
        for c, lbl in extra_groep[1]:
            header_cell(ws, 3, c, lbl, kleur_mid)
    else:
        mc(e_in, e_uit); header_cell(ws, 2, e_in, "In / Uit", kleur_mid)
        header_cell(ws, 3, e_in, "In", kleur_mid)
        header_cell(ws, 3, e_uit, "Uit", kleur_mid)
        mc(e_opv1, e_opv2); header_cell(ws, 2, e_opv1, "Opvang", kleur_mid)
        header_cell(ws, 3, e_opv1, "Ochtend/Avond", kleur_mid)
        header_cell(ws, 3, e_opv2, "Middag", kleur_mid)

    mc(dag0, dag0 + n - 1)
    header_cell(ws, 2, dag0, "Aanwezigheid per dagdeel", kleur_mid)
    for k, lbl in enumerate(daglabels):
        header_cell(ws, 3, dag0 + k, lbl, kleur_mid)

    ws.merge_cells(f"{get_column_letter(c_opm)}2:{get_column_letter(c_opm)}3")
    header_cell(ws, 2, c_opm, "Opmerking", kleur_dark)

    vul_rijen(ws, namen, eerste_rij=4, laatste_kol=laatste)

    ws.column_dimensions[get_column_letter(c_num)].width = 4
    ws.column_dimensions[get_column_letter(c_naam)].width = 24
    for c in range(3, laatste):
        ws.column_dimensions[get_column_letter(c)].width = 8
    ws.column_dimensions[get_column_letter(c_opm)].width = 18


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Combineer twee Kwandoo-aanwezigheidslijsten tot één Excel.")
    ap.add_argument("pdf1", help="Eerste PDF (Jadadde of Villa Max).")
    ap.add_argument("pdf2", help="Tweede PDF (de andere lijst).")
    ap.add_argument("-o", "--output", default="Gecombineerde_aanwezigheidslijst.xlsx",
                    help="Pad voor het Excel-bestand.")
    args = ap.parse_args()

    lijsten = {}
    datums_per = {}
    for pad in (args.pdf1, args.pdf2):
        tekst = lees_pdf_tekst(pad)
        soort = detecteer_type(tekst)
        if soort is None:
            sys.exit(f"FOUT: kan niet bepalen of dit Jadadde of Villa Max is: {pad}")
        if soort in lijsten:
            sys.exit(f"FOUT: twee keer een '{soort}'-lijst aangeleverd. "
                     "Geef één Jadadde- en één Villa Max-PDF.")
        lijsten[soort] = extract_namen(tekst)
        datums_per[soort] = parse_datums(tekst)

    if "jadadde" not in lijsten or "villamax" not in lijsten:
        sys.exit("FOUT: er is zowel een Jadadde- als een Villa Max-PDF nodig.")

    jadadde = lijsten["jadadde"]
    villamax = lijsten["villamax"]
    datums = datums_per["jadadde"] or datums_per["villamax"]

    jad_keys = {match_sleutel(n): n for n in jadadde}
    vil_keys = {match_sleutel(n): n for n in villamax}
    overlap = set(jad_keys) & set(vil_keys)

    overlap_namen = [jad_keys[k] for k in overlap]
    alleen_jadadde = [n for n in jadadde if match_sleutel(n) not in overlap]
    alleen_villamax = [n for n in villamax if match_sleutel(n) not in overlap]

    print(f"Jadadde:   {len(jadadde)} namen")
    print(f"Villa Max: {len(villamax)} namen")
    print(f"Overlap:   {len(overlap)}")
    for naam in sorted(overlap_namen, key=sorteer_sleutel):
        print(f"   - {naam}")

    wb = Workbook()
    bouw_tab_gecombineerd(wb, overlap_namen, datums)
    bouw_tab_alleen(wb, "Alleen Jadadde", alleen_jadadde, datums,
                    BLAUW_DARK, BLAUW_MID, met_maaltijd=False)
    bouw_tab_alleen(wb, "Alleen Villa Max", alleen_villamax, datums,
                    GROEN_DARK, GROEN_MID, met_maaltijd=True)

    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(args.output)
    print(f"\nOpgeslagen: {args.output}")


if __name__ == "__main__":
    main()
